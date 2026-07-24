import datetime as dt
import json
import os
from collections import defaultdict
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK_PATH = Path(os.environ.get("DASHBOARD_WORKBOOK", ROOT / "source-dashboard-latest.xlsx"))
OUTPUT_PATH = Path(__file__).resolve().parents[1] / "public" / "data" / "dashboard.json"


def number(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def date_key(value):
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return None


def header_index(row, *labels):
    normalized = [str(value).replace(" ", "").lower() if value is not None else "" for value in row]
    for label in labels:
        target = label.replace(" ", "").lower()
        for index, value in enumerate(normalized):
            if value == target:
                return index
    raise ValueError(f"Missing required column: {labels}")


workbook = openpyxl.load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
daily = defaultdict(lambda: {"shopee": 0.0, "tiktok": 0.0, "lazada": 0.0, "shopeeSubsidy": 0.0, "tiktokSubsidy": 0.0})

shopee_sheet = workbook["DMS shopee店铺实收GMV"]
shopee_rows = shopee_sheet.iter_rows(values_only=True)
shopee_header = next(shopee_rows)
shopee_date_col = header_index(shopee_header, "Order Date", "日期")
shopee_gmv_col = header_index(shopee_header, "实收GMV（人民币）", "实收GMV(人民币）", "实收GMV(人民币)")
shopee_subsidy_col = header_index(shopee_header, "平台补贴")

for row in shopee_rows:
    day = date_key(row[shopee_date_col] if len(row) > shopee_date_col else None)
    if not day:
        continue
    daily[day]["shopee"] += number(row[shopee_gmv_col] if len(row) > shopee_gmv_col else 0)
    daily[day]["shopeeSubsidy"] += number(row[shopee_subsidy_col] if len(row) > shopee_subsidy_col else 0)

for row in workbook["DMS TIKTOK 店铺实收gmv"].iter_rows(min_row=2, values_only=True):
    day = date_key(row[8] if len(row) > 8 else None)
    if not day:
        continue
    daily[day]["tiktok"] += number(row[9] if len(row) > 9 else 0)
    daily[day]["tiktokSubsidy"] += number(row[10] if len(row) > 10 else 0)

for row in workbook["G2G 【Lazada】销售数据源"].iter_rows(min_row=2, values_only=True):
    day = date_key(row[1] if len(row) > 1 else None)
    if not day:
        continue
    daily[day]["lazada"] += number(row[19] if len(row) > 19 else 0)

targets = {}
exchange_rate = None
for row in workbook["月度目标"].iter_rows(min_row=2, values_only=True):
    label = row[0] if row else None
    if not label:
        continue
    text = str(label)
    year = 2000 + int(text[:2])
    month = int(text.split("年", 1)[1].split("月", 1)[0])
    month_key = f"{year:04d}-{month:02d}"
    targets[month_key] = {
        "shelf": number(row[1] if len(row) > 1 else 0),
        "tiktok": number(row[2] if len(row) > 2 else 0),
    }
    if len(row) > 4 and number(row[4]):
        exchange_rate = number(row[4])

rows = []
for day, values in sorted(daily.items()):
    if not any(values.values()):
        continue
    shelf = values["shopee"] + values["lazada"]
    total = shelf + values["tiktok"]
    subsidy = values["shopeeSubsidy"] + values["tiktokSubsidy"]
    rows.append({
        "date": day,
        "country": "新加坡",
        "brand": "Glad2Glow",
        **{key: round(value, 2) for key, value in values.items()},
        "shelf": round(shelf, 2),
        "total": round(total, 2),
        "subsidy": round(subsidy, 2),
    })

as_of = max((row["date"] for row in rows), default=None)
payload = {
    "meta": {
        "asOf": as_of,
        "currency": "CNY",
        "country": "新加坡",
        "brand": "Glad2Glow",
        "source": "Google Sheets snapshot",
        "lazadaSubsidyIncluded": False,
        "exchangeRate": exchange_rate,
    },
    "targets": targets,
    "daily": rows,
}

OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
print(f"wrote {OUTPUT_PATH} with {len(rows)} daily rows")
